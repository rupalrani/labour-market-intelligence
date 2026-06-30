from openpyxl import load_workbook
wb=load_workbook("Project4_Labour_Workbook.xlsx", data_only=True)
o=[]
s=wb["11_Final_Summary"]
o.append("Summary C10-C14: %s %s %s %s %s"%(s["C10"].value,s["C11"].value,s["C12"].value,s["C13"].value,s["C14"].value))
o.append("Summary refs total/live/incl C16-C18: %s %s %s"%(s["C16"].value,s["C17"].value,s["C18"].value))
pr=wb["06_PRISMA"]
o.append("PRISMA B11 incl=%s"%pr["B11"].value)
for r in range(10,16):
    v=pr.cell(row=r,column=2).value
    if v in ("MATCH","CHECK"): o.append("PRISMA consistency=%s"%v)
rf=wb["10_References"]
for r in range(14,18):
    a=rf.cell(row=r,column=1).value; c=rf.cell(row=r,column=3).value
    if a: o.append("REF %s = %s"%(a,c))
open("v4.txt","w").write("\n".join(str(x) for x in o))
